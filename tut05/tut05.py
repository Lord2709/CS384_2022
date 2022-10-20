#Help https://youtu.be/N6PBd4XdnEw
import os
os.system('cls')
import openpyxl
import math
import numpy as np

from datetime import datetime
start_time = datetime.now()

# write_in_xlsx function is used to write values from list to the xlsx file created
def write_in_xlsx(range_of_rows,column_number,ls,ws):
    for i in range(1,range_of_rows+1):
        cellref=ws.cell(row=i, column=column_number)
        if i == 1:
            cellref.value = ls[0]
        else:
            cellref.value=ls[i-1]

# calculate_rank function is used to calculate rank to map the ranks
def calculate_rank(vector):
    a={}
    rank=1
    for num in sorted(vector, reverse=True):
        if num not in a:
            a[num]=rank
            rank=rank+1
    return[a[i] for i in vector]
    
def octant_range_names(mod=5000):
    try:
        if os.path.exists("octant_output_ranking_excel.xlsx"):
            try:
                os.remove("octant_output_ranking_excel.xlsx")
            except PermissionError:
                print("octant_output_ranking_excel.xlsx is being used !!!")

        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
        
        wb = openpyxl.load_workbook('octant_input.xlsx')
        ws = wb.active
        rows = ws.iter_rows(min_row = ws.min_row, max_row=ws.max_row, min_col = ws.min_column, max_col = ws.max_column)
        # print(rows)

        # New workbook created to store the calculated values and save it as octant_output_ranking_excel
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        mr = ws.max_row
        mc = ws.max_column
        # Below code is used to copy the input file data to the output file
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                c = ws.cell(row = i, column = j)
                ws_output.cell(row = i+1, column = j).value = c.value
        

        output = [[],[],[],[]]
        l_col = ['A','B','C','D']
        for i in range(4):
            col = ws_output[l_col[i]][2:]
            for cell in col:
                output[i].append(cell.value)
        # print(ws_output['A29747'].value)

        ws_output['E1'].value = " "
        ws_output['E2'].value = "U Avg"
        ws_output['E3'].value = np.mean(output[1])
        ws_output['F1'].value = " "
        ws_output['F2'].value = "V Avg"
        ws_output['F3'].value = np.mean(output[2])
        ws_output['G1'].value = " "
        ws_output['G2'].value = "W Avg"
        ws_output['G3'].value = np.mean(output[3])


        H = output[1] - ws_output['E3'].value
        I = output[2] - ws_output['F3'].value
        J = output[3] - ws_output['G3'].value
        H_col = [" ","U'=U-U Avg"]
        I_col = [" ","V'=V-V Avg"]
        J_col = [" ","W'=W-W Avg"]
        for i in range(len(H)):
            H_col.append(H[i])
            I_col.append(I[i])
            J_col.append(J[i])
        # print(H_col[29745])
        len_H = len(H_col)
        write_in_xlsx(len_H, 8, H_col, ws_output)
        write_in_xlsx(len_H, 9, I_col, ws_output)
        write_in_xlsx(len_H, 10, J_col, ws_output)
        # Above code is used to store values in output file for H,I,J columns respectively

        Oct = [" ","Octant"]
        for i in range(len_H):
            if i==0 or i==1: 
                continue
            else:
                if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(1)
                if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-1)
                if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(2)
                if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-2)
                if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(3)
                if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-3)
                if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(4)
                if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-4)
        write_in_xlsx(len_H, 11, Oct, ws_output)

        L_col = ['','','','User Input']
        write_in_xlsx(len(L_col), 12, L_col, ws_output)

        M_col = ["", "Octant ID", "Overall Count",f"Mod {mod}"]

        n = len(Oct[2:])
        if n%mod == 0:
            n_ranges = n//mod
        else:
            n_ranges = math.ceil(n/mod)

        # Below code is used to give ranges to the overall count table
        t = 0
        u = t + mod
        for j in range(n_ranges):
            if u <= n:
                if t==0:
                    M_col.append(f"0000 - {u-1}")
                else:
                    M_col.append(f"{t} - {u-1}")
            else:
                M_col.append(f"{t} - {len_H - 2}")
            t = u
            u += mod
        write_in_xlsx(len(M_col), 13, M_col, ws_output)

        # 2D list (overall_count) is used to store the count values of {+1,-2,+2,-2,+3,-3,+4,-4} for entire dataset and respective mod ranges
        overall_count = []
        e = 0
        st = ["+1","-1","+2","-2","+3","-3","+4","-4"]
        for i in range(8):
            empty_list = [""]
            empty_list.append(st[e])
            overall_count.append(empty_list)
            e += 1

        df2 = Oct[2:]
        int_l = [1,-1,2,-2,3,-3,4,-4]
        j = 0
        for i in range(8):
            overall_count[i].append(df2.count(int_l[j]))
            overall_count[i].append(" ")
            j += 1
        

        N_col = ["","","","Octant ID"] + int_l
        O_col = ["","","","Octant Name","Internal outward interaction","External outward interaction",
                "External Ejection","Internal Ejection","External inward interaction",
                "Internal inward interaction","Internal sweep","External sweep"]

        # Below while function is used to count the {+1,-2,+2,-2,+3,-3,+4,-4} values for respective mod ranges
        k = 0
        m = mod
        n_r = n_ranges
        while n_r > 0:   
            n_r -= 1
            d_f = df2[k:m]
            j = 0
            for i in range(8):
                overall_count[i].append(d_f.count(int_l[j]))
                j += 1
            k = m 
            m = m + mod

        overall_count[0] += N_col
        overall_count[1] += O_col

        # print(overall_count)
        col = 14
        for i in range(8):
            if i == 2:
                col += 1
                continue
            write_in_xlsx(len(overall_count[i]), col, overall_count[i], ws_output)
            col += 1
        # Above code is used to store values in output file from M to U columns respectively except Q column
        # wb_output.save("octant_output_ranking_excel.xlsx")

        # overall_rank 2d list is created to store all the columns containing rank values of respective mod ranges
        overall_rank = [[1,"Rank of 1"],[-1,"Rank of 2"],[2,"Rank of 3"],[-2,"Rank of 4"],[3,"Rank of 5"],[-3,"Rank of 6"],[4,"Rank of 7"],[-4,"Rank of 8"]]

        row_list = []
        imp_element_pos = [2] + [i for i in range(4,n_ranges+4)]
        for i in imp_element_pos:
            l = []
            for j in range(8):
                l.append(overall_count[j][i])
            row_list.append(l)


        rank_list = []
        for i in range(len(row_list)):
            rank_list.append(calculate_rank(row_list[i]))

        for i in range(len(overall_rank)):
            for j in rank_list:
                overall_rank[i].append(j[i])
            overall_rank[i].insert(3,"")
            # print(overall_rank[1])

        col = 22
        for i in range(8):
            write_in_xlsx(len(overall_rank[i]), col, overall_rank[i], ws_output)
            col += 1
        # Above code is used to store values in output file from V to AC columns respectively 

        
        AD_col = ["","Rank1 Octant ID"]
        for i in rank_list:
            x = i.index(min(i))
            AD_col.append(int_l[x])

        AE_col = ["","Rank1 Octant Name"]
        for i in AD_col[2:]:
            AE_col.append(octant_name_id_mapping[str(i)])
        # print(AE_col)

        AE_col_copy = list(AE_col[3:])
        # print(AE_col_copy)
        AD_col.insert(3,"")
        AE_col.insert(3,"")
        write_in_xlsx(len(AD_col),30,AD_col,ws_output)
        write_in_xlsx(len(AE_col),31,AE_col,ws_output)
        # Above code is used to store values in output file from AD to AE columns respectively 

        final_mapping = {"Internal outward interaction":0, "External outward interaction":0, "External Ejection":0, "Internal Ejection":0, "External inward interaction":0, "Internal inward interaction":0, "Internal sweep":0, "External sweep":0}
        for item in AE_col_copy:
            if (item in final_mapping):
                final_mapping[item] += 1
            else:
                final_mapping[item] = 1

        # print(final_mapping)
        final = []
        for i in O_col[4:]:
            final.append(final_mapping[i])

        # print(final)
        Q_col = ["","","","Count of Rank 1 Mod Values"] + final
        overall_count[2] += Q_col
        write_in_xlsx(len(overall_count[2]),16,overall_count[2],ws_output)

        try:
            wb_output.save("octant_output_ranking_excel.xlsx")
            print("Code Compiled Successfully")
        except:
            print("Error : It seems that no output file has been generated.")
    except:
        print("Error in calling function")

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_range_names(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))