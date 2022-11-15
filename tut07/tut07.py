import openpyxl
import os
os.system('cls')
import math
import numpy as np
from openpyxl.styles import Alignment,Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from datetime import datetime
start_time = datetime.now()


current_dir = os.getcwd()

global input_path
input_path = current_dir.replace('\\','/') + "/input/"

# os.chdir(current_dir)
if os.path.isdir("output"):
	list_of_files = os.listdir('output')
	for files in list_of_files:
		os.remove(current_dir+"/output/" + f"{files}")
	os.rmdir("output")

global output_path
output_path = current_dir.replace('\\','/') + "/output/"
os.mkdir(output_path)


def write_in_xlsx(range_of_rows,column_number,ls,ws):
    for i in range(1,range_of_rows+1):
        cellref=ws.cell(row=i, column=column_number)
        if i == 1:
            cellref.value = ls[0]
        else:
            cellref.value=ls[i-1]


def calculate_rank(vector):
    a={}
    rank=1
    for num in sorted(vector, reverse=True):
        if num not in a:
            a[num]=rank
            rank=rank+1
    return[a[i] for i in vector]


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def bgcolor(range,ws):
	fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
	ws.conditional_formatting.add(range, CellIsRule(operator='equal', formula=[1], fill=fill))

def max_color(range1,ws):
	fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

	max_list = []

	for row in ws[range1]:
		list_row_val = []
		for cell in row:
			list_row_val.append(cell.value)
		max_list.append(max(list_row_val))

	for row in ws[range1]:
		for cell in row:
			if cell.value == max_list[0]:
				cell.fill = fill
				if len(max_list) != 0:
					del max_list[0]
				break
			else:
				continue

#Help
def octant_analysis(mod=5000):
	files = os.listdir(input_path)

	octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
	
	for f in files:
		file_path = input_path +"/" + f
		wb = openpyxl.load_workbook(file_path)
		ws = wb.active
		rows = ws.iter_rows(min_row = 0, max_row=29746, min_col = 1, max_col = 21)

		os.chdir(output_path)

		wb_output = openpyxl.Workbook()
		ws_output = wb_output.active
		mr = ws.max_row
		mc = ws.max_column
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
				c = ws.cell(row = i, column = j)
				ws_output.cell(row = i+1, column = j).value = c.value
		

		output = [[],[],[],[]]
		l_col = ['A','B','C','D']
		for i in range(4):
			col = ws_output[l_col[i]][2:]
			for cell in col:
				output[i].append(cell.value)


		ws_output['E1'].value = " "
		ws_output['E2'].value = "U Avg"
		ws_output['E3'].value = np.round_(np.mean(output[1]), decimals = 3)
		ws_output['F1'].value = " "
		ws_output['F2'].value = "V Avg"
		ws_output['F3'].value = np.round_(np.mean(output[2]), decimals = 3)
		ws_output['G1'].value = " "
		ws_output['G2'].value = "W Avg"
		ws_output['G3'].value = np.round_(np.mean(output[3]), decimals = 3)


		H = output[1] - ws_output['E3'].value
		I = output[2] - ws_output['F3'].value
		J = output[3] - ws_output['G3'].value
		H_col = [" ","U'=U-U Avg"]
		I_col = [" ","V'=V-V Avg"]
		J_col = [" ","W'=W-W Avg"]
		for i in range(len(H)):
			H_col.append(np.round_((H[i]), decimals = 3))
			I_col.append(np.round_((I[i]), decimals = 3))
			J_col.append(np.round_((J[i]), decimals = 3))
	

		len_H = len(H_col)
		write_in_xlsx(len_H, 8, H_col, ws_output)
		write_in_xlsx(len_H, 9, I_col, ws_output)
		write_in_xlsx(len_H, 10, J_col, ws_output)


		Oct = [" ","Octant"]
		for i in range(len_H):
			if i==0 or i==1: 
				continue
			else:
				if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(1)
				if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-1)
				if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(2)
				if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-2)
				if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(3)
				if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-3)
				if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(4)
				if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-4)
		write_in_xlsx(len_H, 11, Oct, ws_output)
		
		M_col = ['','','',f'Mod {mod}']
		write_in_xlsx(len(M_col), 13, M_col, ws_output)

		N_col = ["Overall Octant Count","", "Octant ID", "Overall Count"]

		n = len(Oct[2:])
		if n%mod == 0:
			n_ranges = n//mod
		else:
			n_ranges = math.ceil(n/mod)


		t = 0
		u = t + mod
		for j in range(n_ranges):
			if u <= n:
				if t==0:
					N_col.append(f"0000 - {u-1}")
				else:
					N_col.append(f"{t} - {u-1}")
			else:
				N_col.append(f"{t} - {len_H - 2}")
			t = u
			u += mod
		write_in_xlsx(len(N_col), 14, N_col, ws_output)

		# 8 new list are created to store the count value of {+1,-2,+2,-2,+3,-3,+4,-4}
		overall_count = []
		e = 0
		st = ["+1","-1","+2","-2","+3","-3","+4","-4"]
		for i in range(8):
			empty_list = ["",""]
			empty_list.append(st[e])
			overall_count.append(empty_list)
			e += 1

		df2 = Oct[2:]
		int_l = [1,-1,2,-2,3,-3,4,-4]
		j = 0
		for i in range(8):
			overall_count[i].append(df2.count(int_l[j]))
			j += 1
		# n --> given in tut01.pdf that max value will never exceed 30000.
		# Below while function is used to count the {+1,-2,+2,-2,+3,-3,+4,-4} values for different mod's

		k = 0
		m = mod
		n_r = n_ranges
		while n_r > 0:   
			n_r -= 1
			d_f = df2[k:m]
			j = 0
			for i in range(8):
				overall_count[i].append(d_f.count(int_l[j]))
				j += 1
			k = m 
			m = m + mod

		# overall_count[0] += N_col
		# overall_count[1] += O_col

		# print(overall_count)
		col = 15
		for i in range(8):
			# if i == 2:
			# 	col += 1
			# 	continue
			write_in_xlsx(len(overall_count[i]), col, overall_count[i], ws_output)
			col += 1

		overall_rank = [["","","Rank Octant 1"],["","","Rank Octant -1"],["","","Rank Octant 2"],["","","Rank Octant -2"],["","","Rank Octant 3"],["","","Rank Octant -3"],["","","Rank Octant 4"],["","","Rank Octant -4"]]
		AC_col = ["","Octant ID"] + int_l
		AD_col = ["","Octant Name","Internal outward interaction","External outward interaction",
				"External Ejection","Internal Ejection","External inward interaction",
				"Internal inward interaction","Internal sweep","External sweep"]

		row_list = []
		imp_element_pos = [2] + [i for i in range(4,n_ranges+4)]
		for i in imp_element_pos:
			l = []
			for j in range(8):
				l.append(overall_count[j][i])
			row_list.append(l)


		rank_list = []
		for i in range(len(row_list)):
			rank_list.append(calculate_rank(row_list[i]))

		# print(overall_rank)
		for i in range(len(overall_rank)):
			for j in rank_list:
				overall_rank[i].append(j[i])
			# overall_rank[i].insert(3,"")
			# print(overall_rank[1])

		overall_rank[6] += AC_col
		overall_rank[7] += AD_col

		col = 23
		for i in range(8):
			write_in_xlsx(len(overall_rank[i]), col, overall_rank[i], ws_output)
			col += 1
		
		AE_col = ["","","Rank1 Octant ID"]
		for i in rank_list:
			x = i.index(min(i))
			AE_col.append(int_l[x])

		AF_col = ["","","Rank1 Octant Name"]
		for i in AE_col[3:]:
			AF_col.append(octant_name_id_mapping[str(i)])
		# print(AF_col)

		AF_col_copy = list(AF_col[4:])
		# print(AF_col_copy)
		# write_in_xlsx(len(AE_col),31,AE_col,ws_output)
		write_in_xlsx(len(AF_col),32,AF_col,ws_output)
	
		final_mapping = {"Internal outward interaction":0, "External outward interaction":0, "External Ejection":0, "Internal Ejection":0, "External inward interaction":0, "Internal inward interaction":0, "Internal sweep":0, "External sweep":0}
		for item in AF_col_copy:
			if (item in final_mapping):
				final_mapping[item] += 1
			else:
				final_mapping[item] = 1

		# print(final_mapping)
		final = []
		for i in AD_col[2:]:
			final.append(final_mapping[i])

		# print(final)
		AE_col_ = ["","Count of Rank 1 Mod Values"] + final
		AE_col += AE_col_
		write_in_xlsx(len(AE_col),31,AE_col,ws_output)

		# ========================================== Tut05 Part-End ===============================================

		# ========================================== Tut02 Part-Start =============================================

		AH_col = ['','','']
		for i in range(n_ranges+1):
			AH_col.append("From")
			AH_col += ["" for i in range(13)]
		write_in_xlsx(len(AH_col), 34, AH_col, ws_output)


		AI_col = ["Overall Transition","","Octant #"] + st

		q = 0
		r = mod
		nr = n_ranges
		while nr > 0:   
			nr -= 1
			if r <= n:
				if q==0:
					p = f"{0000} - {r-1}"
				else:
					p = f"{q} - {r-1}"
			else:
				p = f"{q} - {len_H - 2}"
			AI_col += ["","","","Mod Transition Count"]
			AI_col.append(p)
			AI_col.append("Octant #")
			AI_col += st
			q = r 
			r = r + mod
		write_in_xlsx(len(AI_col), 35, AI_col, ws_output)


		# 8 new list are created to store the count value of {+1,-2,+2,-2,+3,-3,+4,-4}
		overall_count1 = []
		e = 0
		for i in range(8):
			empty_list = ["",""]
			empty_list.append(st[e])
			overall_count1.append(empty_list)
			e += 1


		for j in range(8):
			l = [0]*8
			d_f = df2
			for i in range(2,len(d_f)+1):  
				if(d_f[i-1] == int_l[j]):
					if(d_f[i-2] == 1):
						l[0] += 1
					elif(d_f[i-2] == -1):
						l[1] += 1
					elif(d_f[i-2] == 2):
						l[2] += 1
					elif(d_f[i-2] == -2):
						l[3] += 1
					elif(d_f[i-2] == 3):
						l[4] += 1
					elif(d_f[i-2] == -3):
						l[5] += 1
					elif(d_f[i-2] == 4):
						l[6] += 1
					elif(d_f[i-2] == -4):
						l[7] += 1
			if j==0:
				l+=["","","","","To"]
			else:
				l += ["" for i in range(5)]

			overall_count1[j] += l  

		q = 0
		r = mod
		n_r_ = n_ranges
		while n_r_>0:
			n_r_ -= 1
			for j in range(8):
				overall_count1[j].append(st[j])
				l = [0]*8
				d_f = df2[q:r+1]
				for i in range(2,len(d_f)+1):  
					if(d_f[i-1] == int_l[j]):
						if(d_f[i-2] == 1):
							l[0] += 1
						elif(d_f[i-2] == -1):
							l[1] += 1
						elif(d_f[i-2] == 2):
							l[2] += 1
						elif(d_f[i-2] == -2):
							l[3] += 1
						elif(d_f[i-2] == 3):
							l[4] += 1
						elif(d_f[i-2] == -3):
							l[5] += 1
						elif(d_f[i-2] == 4):
							l[6] += 1
						elif(d_f[i-2] == -4):
							l[7] += 1
				if j==0 and (n_r_ != 0):
					l+=["","","","","To"]
				else:
					l += ["" for i in range(5)]

				overall_count1[j] += l  
			q = r 
			r += mod

		col = 36
		for i in range(8):
			write_in_xlsx(len(overall_count1[i]), col, overall_count1[i], ws_output)
			col += 1
		ws_output.cell(row=2, column=36).value = "To"

		# ========================================== Tut02 Part-End =================================================

		# ========================================== Tut03 Part-Start ===============================================

		g1 = ["Longest Subsquence Length"," ","Octant ##","+1","-1","+2","-2","+3","-3","+4","-4"]
		g2 = ["","","Longest Subsquence Length"]
		g2_ = []
		g3 = ["","","Count"]
		g3_ = []

		oct = Oct[2:]
		T = []
		for col in ws_output['A']:
			T.append(col.value)
		time1 = T[2:]
		n = len_H - 2

		'''count1 to store initial count and previous_count is used to store 
        previous count as the count value keeps on changing.'''
		for j in int_l:
			count1 = 0
			prev = 0
			indexend = 0
			for i in range(n):
				if(oct[i] == j):
					count1 += 1
				else:
					if(count1 > prev):
						prev = count1
						indexend = i
					count1 = 0
        	# c is used to store total number of times the small longest subsequence occurs
			c = 0
			count2 = 0
			for i in range(n):
				if(oct[i] != j):
					count2 = 0
				else:
					count2 += 1
					if(count2 == prev):
						c+=1
						count2 = 0
            #print(prev, c)
			g2_.append(prev)
			g3_.append(c)
        
        #print(g2_,g3_)
		g2 = g2 + g2_
		g3 = g3 + g3_

        # 2d list to store entire small longest subsequence table
		g_final = []
		g_final.append(g1)
		g_final.append(g2)
		g_final.append(g3)

		col = 45
		for i in range(3):
			write_in_xlsx(len(g_final[i]), col, g_final[i], ws_output)
			col += 1

		# ========================================== Tut03 Part-End ===================================================
		
		# ========================================== Tut04 Part-Start =================================================

		tut_04(n,time1,oct,int_l,g3_,ws_output)

		# ========================================== Tut04 Part-End ===================================================

		# ========================================== Border - Start ==================================================

		set_border(ws_output, f'N3:AF{n_ranges+4}')
		set_border(ws_output, f'AC{n_ranges+6}:AE{n_ranges+14}')
		set_border(ws_output, 'AS3:AU11')

		column_start = 'AI'
		column_end = 'AQ'
		x = 3
		for i in range(n_ranges+1):
			start = column_start + f'{x}'
			end = column_end + f'{x+8}'
			set_border(ws_output, f'{start}:{end}')
			# print(start,end)
			x += 14

		y = sum(g3_) + 19
		set_border(ws_output, f'AW3:AY{y}')
		# ============================================ Border - End ==================================================

		# ============================================ Bgcolor - Start ===============================================

		bgcolor('W4:AD10',ws_output)
		column_start = 'AJ'
		column_end = 'AQ'
		x = 4
		for i in range(n_ranges+1):
			start = column_start + f'{x}'
			end = column_end + f'{x+7}'
			max_color(f'{start}:{end}', ws_output)
			# print(start,end)
			x += 14

		# ============================================ Bgcolor - End =================================================

		output_file_name = f.split(".xlsx")[0] + f"_octant_analysis_mod_{mod}" + ".xlsx"
		wb_output.save(output_file_name)
		print(f"{output_file_name} file compiled successfully")


def tut_04(n,time1,oct,int_l,g3_,ws_output):
	dk = []
	# g0,g1,g2,g3 are the list created which will store the values of the column 
	x = sum(g3_) + 18
	# g0 = [" " for i in range(x)]
	g1 = ["Longest Subsquence Length with Range","","Octant ###"]
	l_ = ["+1","-1","+2","-2","+3","-3","+4","-4"]
	for i in range(len(l_)):
		g1.append(l_[i])
		g1.append("Time")
		gap = [" " for _ in range(g3_[i])]
		g1 = g1 + gap

	g2 = ["","","Longest Subsquence Length"]
	g3 = ["","","Count"]
	k0 = []
	k1 = []
	l_0 = []
	l_1 = []

	for j in int_l:
		count1 = 0
		prev = 0
		for i in range(n):
			if(oct[i] == j):
				count1 += 1
			else:
				if(count1 > prev):
					prev = count1
					indexend = i
				count1 = 0
		c = 0
		count2 = 0
		indexend = 0
		l0 = []
		l1 = []   
		for i in range(n):
			if(oct[i] != j):
				count2 = 0
			else:
				count2 += 1
			if(count2 == prev):
				c+=1
				count2 = 0
				indexend = i
				l0.append(indexend-prev+1)
				l1.append(indexend)
			#print(prev, c)
		l_0.append(l0)
		l_1.append(l1)
		k0.append(prev)
		k1.append(c)
	
	for j in range(8):
		g2.append(k0[j])
		g2.append("From")
		z1 = l_0[j]
		for p in range(len(z1)):
			t = time1[z1[p]]
			g2.append(t)

	for j in range(8):
		g3.append(k1[j])
		g3.append("To")
		z2 = l_1[j]
		for p in range(len(z2)):
			t = time1[z2[p]]
			g3.append(t)
			

	dk.append(g1)
	dk.append(g2)
	dk.append(g3)

	col = 49
	for i in range(3):
		write_in_xlsx(len(dk[i]), col, dk[i], ws_output)
		col += 1


##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
